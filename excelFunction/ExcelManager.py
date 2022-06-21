import csv
import json
import pandas as pd
from main import MainWindow
from tokenFunction.Toke import Token
import openpyxl

listProfile = []


def readFileExcel(pathFile, sheetName):
    df = pd.read_excel(pathFile, sheet_name=sheetName)
    return df


def getListProfile(path: str, sheet: str):
    listData = readFileExcel(path, sheet)
    listSTT = listData["STT"]
    listSeedPhrase = listData["Metamask SeePhrase"]
    listMetamaskAddress = listData["Metamask Address"]
    listMetamaskKey = listData["Metamask Key"]
    listPhantomAddress = listData["Phantom Address"]
    listPhantomKey = listData["Phantom Key"]
    for index, item in enumerate(listSTT):
        token = Token()
        token.STT = listSTT[index]
        token.SeedPhrase = listSeedPhrase[index]
        token.MetamaskAddress = listMetamaskAddress[index]
        token.MetamaskKey = listMetamaskKey[index]
        token.PhantomAddress = listPhantomAddress[index]
        token.MetamaskKey = listPhantomKey[index]
        listProfile.append(token)
    return listProfile


def getListDataExcel(path: str, sheet: str):
    listData = readFileExcel(path, sheet)
    listProfile = listData.to_dict('records')
    return listProfile, len(listProfile)


def readExcel(excelPath):
    try:
        file_location = excelPath
        df = pd.read_excel(file_location, sheet_name="Sheet1")
        return df
    except Exception:
        MainWindow.ShowError("Lỗi mở file", "Thông báo")


def writerCsv(header, data, filename):
    with open(filename, "w", newline="") as csvfile:
        movies = csv.writer(csvfile)
        movies.writerow(header)
        for x in data:
            movies.writerow(x)


def readCsv(header, filename):
    df = pd.read_csv(filename, usecols=header)
    return df


def saveFileJson(filePath, data):
    try:
        with open(filePath, 'w', encoding='utf-8') as file:
            json.dump(data, file)
            file.close()
    except Exception as ex:
        print("Lỗi ghi file json")


def redFileJson(filePath):
    try:
        with open(filePath, 'r', encoding='utf-8') as file:
            data = json.loads(file.read())
            file.close()
            return data
    except Exception as ex:
        print("Lỗi read file json")


def saveScanToken(path, listToken):
    try:
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        column_list = [cell.value for cell in ws[1]]
        count = 1
        while True:
            if count == 1:
                for j in range(len(column_list)):
                    if column_list[j] == "Address":
                        inAddress = j + 1
                    if column_list[j] == "Sym":
                        inSym = j + 1
                    if column_list[j] == "Amount":
                        inAmount = j + 1
                    if column_list[j] == "Contract":
                        inContract = j + 1
                    if column_list[j] == "PrivateKey":
                        inPrivateKey = j + 1
            if count >= 2:
                for items in listToken:
                    ws.cell(count, inAddress).value = items.get("Address")
                    ws.cell(count, inSym).value = items.get("name")
                    ws.cell(count, inAmount).value = items.get("balance")
                    ws.cell(count, inContract).value = items.get("token_address")
                    ws.cell(count, inPrivateKey).value = items.get("Key")
                    count += 1
                wb.save(filename=path)
                wb.close()
                break
            count += 1
    except Exception:
        print("Ghi file lỗi ")